#!/usr/bin/env python3
"""
Crawler ANAF — Buletin Statistic Fiscal (venituri bugetare pe tipuri de impozite).

Descarcă buletinele statistice fiscale trimestriale ANAF de pe data.gov.ro,
parsează capitolul „4.1. Venituri bugetare realizate" (și opțional
„2.2. Contribuabili activi") și indexează în Meilisearch.

Sursă: data.gov.ro — pachete „Buletin statistic fiscal nr. {T} {AN}" (XLSX)

Structură capitol 4.1 (verificat pe nr. 1/2026, openpyxl):
    r130: Denumire buget | Realizări | | | Trim. I 2026 faţă de
    r131:                  | Anul 2025 | Trim. I 2025 | Trim. I 2026 | trim. I 2025
    r132:                  | - milioane lei - | | | - procente -
    r133: Total venituri administrate de ANAF | 511.648,4 | 108.960,9 | 119.228,0 | 109,4
    r134: Buget de stat | 320.445,1 | 63.094,0 | 69.341,3 | 109,9
    ... (TVA, Accize, Impozit pe profit...)
Valori: milioane lei, virgulă zecimală („511.648,4")

Usage:
    python crawler_anaf.py              # Indexează toate buletinele găsite
    python crawler_anaf.py --dry-run    # Doar descarcă, nu indexează
    python crawler_anaf.py --max=2      # Doar primele 2 buletine
"""

import argparse
import json
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path

import openpyxl
import requests
from dotenv import load_dotenv
from meilisearch import Client

# ── Config ─────────────────────────────────────

load_dotenv(Path(__file__).resolve().parent.parent / "frontend" / ".env")

MEILISEARCH_HOST = os.getenv("MEILISEARCH_HOST", "http://localhost:7700")
MEILISEARCH_API_KEY = os.getenv("MEILISEARCH_API_KEY", "")
INDEX_NAME = "taxe_impozite"
BATCH_SIZE = 300
STATE_FILE = Path(__file__).parent / ".crawler_state.json"

CKAN_API = "https://data.gov.ro/api/3/action/package_search"
HEADERS = {
    "User-Agent": "OpenTransparentaBot/1.0 (+https://github.com/Dragos-AI13/open-transparenta; date publice ANAF)",
    "Accept": "application/json",
}

# Trimit doar din 2023 încoace (structură mai uniformă, date relevante)
MIN_YEAR = 2023


def log(msg: str):
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] {msg}", flush=True)


# ── State ──────────────────────────────────────

def load_state() -> dict:
    if STATE_FILE.exists():
        try:
            return json.loads(STATE_FILE.read_text())
        except Exception:
            pass
    return {"anaf_hash": None, "anaf_count": 0}


def save_state(state: dict):
    STATE_FILE.write_text(json.dumps(state, indent=2, default=str))


# ── CKAN discovery ─────────────────────────────

def find_buletine() -> list[tuple[int, int, str]]:
    """Găsește buletinele statistice fiscale → [(trimestru, an, url_xlsx)]."""
    out = []
    page = 0
    while True:
        try:
            resp = requests.get(
                CKAN_API,
                params={
                    "q": "Buletin statistic fiscal",
                    "rows": 50,
                    "start": page * 50,
                    "sort": "metadata_modified desc",
                },
                headers=HEADERS,
                timeout=30,
            )
            resp.raise_for_status()
            results = resp.json()["result"]["results"]
        except Exception as e:
            log(f"  CKAN error: {e}")
            break
        if not results:
            break
        for pkg in results:
            title = pkg.get("title", "")
            m = re.match(r"Buletin statistic fiscal nr\.?\s*(\d+)\s+(\d{4})", title, re.I)
            if not m:
                continue
            trim, an = int(m.group(1)), int(m.group(2))
            if an < MIN_YEAR:
                continue
            url = None
            for res in pkg.get("resources", []):
                fmt = (res.get("format") or "").lower()
                name = (res.get("name") or "").lower()
                if "xlsx" in fmt or "xls" in fmt or name.endswith(".xlsx"):
                    url = res.get("url")
                    break
            if url:
                out.append((trim, an, url))
        page += 1
        if page > 5:  # max ~300 pachete
            break
    # Dedupe + sort desc după (an, trimestru)
    seen = set()
    uniq = []
    for t, a, u in out:
        if (t, a) not in seen:
            seen.add((t, a))
            uniq.append((t, a, u))
    uniq.sort(key=lambda x: (x[1], x[0]), reverse=True)
    return uniq


# ── Parsing ────────────────────────────────────

def parse_num(value) -> float | None:
    """„511.648,4" → 511648.4 | None dacă gol/invalid."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s or s in ("-", "—", "x", "X"):
        return None
    # Elimină punctele de mii, transformă virgula zecimală
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def parse_venituri_bugetare(ws, trim: int, an: int) -> list[dict]:
    """Extrage capitolul 4.1 — venituri bugetare realizate pe tipuri."""
    docs = []
    # Rândul „Total venituri administrate de ANAF" e UNIC în sheet
    # (cuprinsul are puncte suspensive și nu-l conține exact)
    start_row = None
    for i in range(1, ws.max_row + 1):
        cell = ws.cell(row=i, column=1).value
        if cell and "Total venituri administrate" in str(cell):
            start_row = i
            break
    if start_row is None:
        return docs

    for i in range(start_row, ws.max_row + 1):
        indicator = ws.cell(row=i, column=1).value
        if indicator is None or not str(indicator).strip():
            break
        indicator_s = str(indicator).strip()
        # Oprim la nota de subsol (începe cu *)
        if indicator_s.startswith("*") or indicator_s.startswith("Not"):
            break
        # Oprim la următorul capitol (4.2, 5. etc.)
        if re.match(r"^\d+\.\d*\.?\s", indicator_s):
            break

        total_an_anterior = parse_num(ws.cell(row=i, column=2).value)
        t_anterior = parse_num(ws.cell(row=i, column=3).value)
        t_curent = parse_num(ws.cell(row=i, column=4).value)
        indice = parse_num(ws.cell(row=i, column=5).value)

        if t_curent is None and total_an_anterior is None:
            continue

        slug = re.sub(r"[^a-z0-9]+", "-", indicator_s.lower()).strip("-")[:40]
        docs.append({
            "id": f"{slug}_{an}_T{trim}",
            "indicator": indicator_s,
            "sectiune": "Venituri bugetare",
            "an": an,
            "trimestru": trim,
            "valoare_curent": t_curent,          # Trim. curent (milioane lei)
            "valoare_anterior": t_anterior,      # Trim. anul anterior
            "total_an_anterior": total_an_anterior,  # Realizări anul anterior
            "indice": indice,                    # % față de trim. anterior
            "unitate": "milioane lei",
        })
    return docs


def parse_contribuabili(ws, trim: int, an: int) -> list[dict]:
    """Extrage capitolul 2.2 — contribuabili activi pe categorii."""
    docs = []
    # Caut direct header-ul „Total contribuabili activi" (capitolul 2.2, UNIC)
    # — atenție: „Total contribuabili înregistraţi" (2.1) are altă structură
    start_row = None
    for i in range(1, ws.max_row + 1):
        cell_b = ws.cell(row=i, column=2).value
        if cell_b and "Total contribuabili activi" in str(cell_b):
            start_row = i
            break
    if start_row is None:
        return docs

    # Sar peste header (rândul cu 2025|2026) — caut primul rând cu text în col. A
    for i in range(start_row + 1, min(start_row + 5, ws.max_row + 1)):
        cell_a = ws.cell(row=i, column=1).value
        if cell_a and str(cell_a).strip():
            start_row = i
            break
    else:
        return docs

    for i in range(start_row, ws.max_row + 1):
        indicator = ws.cell(row=i, column=1).value
        if indicator is None or not str(indicator).strip():
            break
        indicator_s = str(indicator).strip()
        if indicator_s.startswith("*") or re.match(r"^\d+\.\d*\.?\s", indicator_s):
            break
        val_an_curent = parse_num(ws.cell(row=i, column=2).value)
        val_an_anterior = parse_num(ws.cell(row=i, column=3).value)
        if val_an_curent is None and val_an_anterior is None:
            continue
        slug = re.sub(r"[^a-z0-9]+", "-", indicator_s.lower()).strip("-")[:40]
        docs.append({
            "id": f"contrib_{slug}_{an}_T{trim}",
            "indicator": indicator_s,
            "sectiune": "Contribuabili înregistrați",
            "an": an,
            "trimestru": trim,
            "valoare_curent": val_an_curent,
            "valoare_anterior": val_an_anterior,
            "total_an_anterior": None,
            "indice": None,
            "unitate": "număr contribuabili",
        })
    return docs


def parse_buletin(content: bytes, trim: int, an: int) -> list[dict]:
    """Parsează un XLSX buletin → documente Meilisearch."""
    from io import BytesIO
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    ws = wb.active
    docs = parse_venituri_bugetare(ws, trim, an)
    docs += parse_contribuabili(ws, trim, an)
    return docs


# ── Meilisearch ────────────────────────────────

STAGING_INDEX = INDEX_NAME + "_staging"


def get_ms_client() -> Client:
    return Client(url=MEILISEARCH_HOST, api_key=MEILISEARCH_API_KEY)


def prepare_staging_index(client: Client):
    try:
        client.delete_index(STAGING_INDEX)
        time.sleep(0.5)
    except Exception:
        pass
    client.create_index(STAGING_INDEX, {"primaryKey": "id"})
    time.sleep(1.5)
    idx = client.index(STAGING_INDEX)
    idx.update_settings({
        "searchableAttributes": ["indicator"],
        "filterableAttributes": ["sectiune", "an", "trimestru"],
        "sortableAttributes": ["an", "valoare_curent"],
    })
    time.sleep(1)
    try:
        client.index(INDEX_NAME).get_stats()
    except Exception:
        log("Indexul live nu există. Îl creez...")
        client.create_index(INDEX_NAME, {"primaryKey": "id"})
        time.sleep(1)
        client.index(INDEX_NAME).update_settings({
            "searchableAttributes": ["indicator"],
            "filterableAttributes": ["sectiune", "an", "trimestru"],
            "sortableAttributes": ["an", "valoare_curent"],
        })
        time.sleep(1)


def swap_indexes(client: Client):
    client.swap_indexes([{"indexes": [INDEX_NAME, STAGING_INDEX]}])
    time.sleep(1)
    try:
        client.delete_index(STAGING_INDEX)
    except Exception:
        pass
    log("Swap atomic complet: staging → live")


# ── Main ───────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--max", type=int, default=None, help="Max buletine de procesat")
    args = parser.parse_args()

    log("Găsesc buletinele statistice fiscale...")
    buletine = find_buletine()
    log(f"Găsite: {len(buletine)} buletine (din {MIN_YEAR}+)")
    if args.max:
        buletine = buletine[: args.max]

    all_docs = []
    for trim, an, url in buletine:
        log(f"  Descarc BSF nr. {trim}/{an}...")
        try:
            resp = requests.get(url, headers=HEADERS, timeout=60)
            if resp.status_code != 200:
                log(f"    HTTP {resp.status_code} — sar peste")
                continue
        except Exception as e:
            log(f"    Eroare: {e} — sar peste")
            continue
        docs = parse_buletin(resp.content, trim, an)
        if docs:
            log(f"    {len(docs)} indicatori extrași")
            all_docs.extend(docs)
        else:
            log("    0 indicatori (structură diferită?)")
        time.sleep(0.8)

    log(f"Total documente: {len(all_docs)}")
    if not all_docs:
        log("Nicio dată extrasă. Ies.")
        sys.exit(1)

    client = get_ms_client()
    prepare_staging_index(client)
    target = client.index(STAGING_INDEX)

    t0 = time.time()
    for i in range(0, len(all_docs), BATCH_SIZE):
        batch = all_docs[i:i + BATCH_SIZE]
        if not args.dry_run:
            task = target.add_documents(batch)
            try:
                client.wait_for_task(task.task_uid, timeout_in_ms=30000)
            except Exception:
                time.sleep(2)
    log(f"Indexat în {time.time() - t0:.1f}s")

    if not args.dry_run:
        swap_indexes(client)
        save_state({"anaf_hash": datetime.now().strftime("%Y%m%d"), "anaf_count": len(all_docs)})
        log("State salvat")
    else:
        log("Dry-run: nu am indexat nimic")


if __name__ == "__main__":
    main()
