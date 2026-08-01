#!/usr/bin/env python3
"""
Crawler MFE — Proiecte contractate din fonduri europene + Stadiul absorbției.

Descarcă de pe data.gov.ro (org: mfe):
  1. 'Proiecte contractate" — XLSX pe 7 programe (POIM, POC, POCU, POR, POAT,
     POAD, POCA), istoric acumulat 2018→prezent. Crawler-ul ia cea mai recentă
     resursă per program.
  2. 'Stadiul absorbtiei fondurilor europene 2014-2020" — XLSX lunar.
  3. 'Stadiul absorbției fonduri europene - Politica de Coeziune 2021-2027" —
     XLSX lunar.

Structură 'Proiecte contractate" (verificat POIM 31 aug 2025, 5.2MB, openpyxl):
    r5:  header românesc (Nr. crt., Titlu proiect, cod SMIS, Nume beneficiar,
         Regiune, Județ, Valoare totala eligibila, Stadiu proiect, PLATI...)
    r7:  header englezesc (duplicat — ignorat)
    r10: primele date reale
    Coloanele 16-19, 24-57: informații suplimentare (cofinanțare, cheltuieli,
    plan financiar pe ani) — nu sunt incluse în index.

Structură 'Stadiul absorbției" (verificat 29 mai 2026):
    sheet 'Fara subtotaluri", header multi-nivel r6-r8, date din r9:
    col4 = nume program, col5 = alocare (EUR), col6 = plăți (EUR),
    col7 = % absorbție (formulă), col8 = prefinanțări (EUR)

Usage:
    python crawler_mfe.py              # Indexează tot
    python crawler_mfe.py --dry-run    # Doar descarcă+parsează, nu indexează
    python crawler_mfe.py --max=2      # Doar 2 programe / 2 resurse absorbție
    python crawler_mfe.py --force      # Ignoră state-ul (re-procesează tot)
"""

import argparse
import hashlib
import json
import os
import re
import sys
import time
from datetime import datetime
from io import BytesIO
from pathlib import Path

import openpyxl
import requests
from dotenv import load_dotenv
from meilisearch import Client

# ── Config ─────────────────────────────────────

load_dotenv(Path(__file__).resolve().parent.parent / "frontend" / ".env")

MEILISEARCH_HOST = os.getenv("MEILISEARCH_HOST", "http://localhost:7700")
MEILISEARCH_API_KEY = os.getenv("MEILISEARCH_API_KEY", "")
INDEX_NAME = "proiecte_fonduri"
INDEX_ABSORBTIE = "absorbtie_fonduri"
BATCH_SIZE = 300
STATE_FILE = Path(__file__).parent / ".crawler_state.json"

CKAN_API = "https://data.gov.ro/api/3/action/package_search"
HEADERS = {
    "User-Agent": "OpenTransparentaBot/1.0 (+https://github.com/Dragos-AI13/open-transparenta; date publice MFE)",
    "Accept": "application/json",
}

# Programele operaționale (prefice în numele resurselor)
PROGRAMS = ["POIM", "POC", "POCU", "POR", "POAT", "POAD", "POCA"]


def log(msg: str):
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] {msg}", flush=True)


# ── Utilitare ───────────────────────────────────

def normalize_ro(s: str) -> str:
    """Normalizează diacriticele românești (ț/ș cu virgulă vs cedilă)."""
    return (
        s.replace("\u015f", "\u0219")  # ș (cedilă → virgulă)
        .replace("\u0163", "\u021b")  # ț
        .replace("\u015e", "\u0218")
        .replace("\u0162", "\u021a")
    )


def parse_num(value) -> float | None:
    """'8037477097.98" sau '1.234.567,89" → float | None."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s or s in ("-", "—", "x", "X"):
        return None
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def doc_hash(*parts: str) -> str:
    return hashlib.sha1("|".join(parts).encode()).hexdigest()[:16]


# ── State ──────────────────────────────────────

def load_state() -> dict:
    if STATE_FILE.exists():
        try:
            return json.loads(STATE_FILE.read_text())
        except Exception:
            pass
    return {"mfe_projects_hash": None, "mfe_projects_count": 0, "mfe_absorb_hash": None, "mfe_absorb_count": 0}


def save_state(state: dict):
    STATE_FILE.write_text(json.dumps(state, indent=2, default=str))


# ── CKAN discovery ─────────────────────────────

def find_package(title_q: str) -> dict | None:
    """Caută primul pachet CKAN care conține query-ul în titlu."""
    try:
        resp = requests.get(
            CKAN_API,
            params={"q": title_q, "rows": 5, "sort": "metadata_modified desc"},
            headers=HEADERS,
            timeout=30,
        )
        resp.raise_for_status()
        for pkg in resp.json()["result"]["results"]:
            if title_q.lower() in pkg.get("title", "").lower():
                return pkg
    except Exception as e:
        log(f"  CKAN error ({title_q}): {e}")
    return None


def find_projects_package() -> dict | None:
    return find_package("Proiecte contractate")


def find_absorb_packages() -> list[dict]:
    """Ambele pachete de absorbție (2014-2020 + 2021-2027)."""
    return [
        p for p in [
            find_package("Stadiul absorbtiei fondurilor europene 2014-2020"),
            find_package("Stadiul absorbției fonduri europene - Politica de Coeziune 2021-2027"),
        ] if p
    ]


def latest_resource_per_program(pkg: dict) -> dict[str, dict]:
    """
    Din pachetul 'Proiecte contractate": cea mai recentă resursă per program.
    Programul = prefixul din numele resursei ('POIM - ...").
    """
    out: dict[str, dict] = {}
    for res in pkg.get("resources", []):
        name = res.get("name", "")
        fmt = (res.get("format") or "").lower()
        if "xlsx" not in fmt and not name.lower().endswith(".xlsx"):
            continue
        prog = None
        for p in PROGRAMS:
            if re.match(rf"^{p}\b", name.upper()):
                prog = p
                break
        if not prog:
            continue
        # 'la 31 august 2025" → dată de raportare pentru comparare
        m = re.search(r"(\d{1,2})\s+(ianuarie|februarie|martie|aprilie|mai|iunie|iulie|august|septembrie|octombrie|noiembrie|decembrie)\s+(\d{4})", name, re.I)
        report_date = m.group(0) if m else ""
        existing = out.get(prog)
        if existing is None or report_date > existing.get("_report_date", ""):
            out[prog] = {**res, "_report_date": report_date}
    return out


# ── Parsing proiecte ───────────────────────────

def find_header_row(ws) -> int:
    """Header-ul românesc: primul rând care conține 'Titlu proiect"."""
    for r in range(1, 12):
        for c in range(1, 25):
            v = ws.cell(row=r, column=c).value
            if v and "Titlu proiect" in str(v):
                return r
    return 5  # fallback verificat


def parse_projects(content: bytes, program: str, report_date: str) -> list[dict]:
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    ws = wb.active
    hdr = find_header_row(ws)

    # Mapare coloane după nume (robust la variații de header)
    col = {}
    for c in range(1, 30):
        v = ws.cell(row=hdr, column=c).value
        if not v:
            continue
        key = str(v).strip().lower()
        if "titlu proiect" in key:
            col["titlu"] = c
        elif "smis" in key:
            col["smis"] = c
        elif "beneficiar" in key:
            col["beneficiar"] = c
        elif "județ" in key or "judet" in key:
            col["judet"] = c
        elif "regiune" in key:
            col["regiune"] = c
        elif "stadiu proiect" in key:
            col["stadiu"] = c
        elif "total valoare proiect" in key:
            col["valoare_totala"] = c
        elif "valoare totala eligibila" in key and "nerambursabila" not in key:
            col["valoare_eligibila"] = c
        elif "plati" in key and "lei" in key:
            col["plati"] = c
        elif "cofinanțare ue" in key or "cofinantare ue" in key:
            col["cofinantare"] = c

    if "titlu" not in col:
        return []

    docs = []
    # Datele încep la hdr+5 (r10 pentru POIM) — dar căutăm primul rând cu titlu
    start = hdr + 5
    for r in range(start, ws.max_row + 1):
        titlu = ws.cell(row=r, column=col["titlu"]).value
        if not titlu or not str(titlu).strip():
            continue
        titlu_s = normalize_ro(str(titlu).strip())
        # Sărim header-ul englezesc ('Project title") dacă apare
        if titlu_s.lower() in ("project title", "titlu proiect"):
            continue

        smis = ws.cell(row=r, column=col.get("smis", 1)).value
        smis_s = str(smis).strip() if smis is not None else ""
        beneficiar = normalize_ro(str(ws.cell(row=r, column=col.get("beneficiar", 1)).value or "").strip())
        judet = normalize_ro(str(ws.cell(row=r, column=col.get("judet", 1)).value or "").strip())
        regiune = normalize_ro(str(ws.cell(row=r, column=col.get("regiune", 1)).value or "").strip())
        stadiu = normalize_ro(str(ws.cell(row=r, column=col.get("stadiu", 1)).value or "").strip())
        valoare_totala = parse_num(ws.cell(row=r, column=col.get("valoare_totala", 1)).value)
        valoare_eligibila = parse_num(ws.cell(row=r, column=col.get("valoare_eligibila", 1)).value)
        plati = parse_num(ws.cell(row=r, column=col.get("plati", 1)).value)
        cofinantare = ws.cell(row=r, column=col.get("cofinantare", 1)).value

        # id: {program}_{smis}_{hash(titlu|judet|valoare)} — SMIS poate lipsi/repeta,
        # iar un proiect poate avea defalcări multiple (valori diferite, ex. SMIS 125325)
        pid_parts = f"{titlu_s}|{judet}|{valoare_totala}"
        pid = f"{program}_{smis_s}_{doc_hash(pid_parts)}" if smis_s else f"{program}_n{doc_hash(pid_parts)}"

        docs.append({
            "id": pid,
            "titlu": titlu_s,
            "program": program,
            "smis": smis_s or None,
            "beneficiar": beneficiar or None,
            "judet": judet or None,
            "regiune": regiune or None,
            "stadiu": stadiu or None,
            "valoare_totala": valoare_totala,
            "valoare_eligibila": valoare_eligibila,
            "plati": plati,
            "cofinantare_ue": cofinantare,
            "data_raportare": report_date,
        })
    return docs


# ── Parsing absorbție ──────────────────────────

def parse_absorb(content: bytes, perioada: str) -> list[dict]:
    """Sheet 'Fara subtotaluri": col4=program, col5=alocare, col6=plăți, col8=prefinanțări."""
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    ws = wb["Fara subtotaluri"] if "Fara subtotaluri" in wb.sheetnames else wb.active
    docs = []
    for r in range(9, ws.max_row + 1):
        nume = ws.cell(row=r, column=4).value
        if not nume or not str(nume).strip():
            continue
        nume_s = normalize_ro(str(nume).strip())
        # Sărim rândurile total (SUM/'TOTAL")
        if nume_s.upper().startswith(("TOTAL", "POLITICA DE COEZIUNE ȘI")):
            continue
        alocare = parse_num(ws.cell(row=r, column=5).value)
        plati = parse_num(ws.cell(row=r, column=6).value)
        prefinantari = parse_num(ws.cell(row=r, column=8).value)
        if alocare is None and plati is None:
            continue
        absorbtie_pct = round((plati / alocare) * 100, 2) if alocare else None
        docs.append({
            "id": doc_hash(perioada, nume_s),
            "program": nume_s,
            "alocare": alocare,
            "plati": plati,
            "absorbtie_pct": absorbtie_pct,
            "prefinantari": prefinantari,
            "perioada": perioada,
            "unitate": "EUR",
        })
    return docs


# ── Meilisearch ────────────────────────────────

def get_ms_client() -> Client:
    return Client(url=MEILISEARCH_HOST, api_key=MEILISEARCH_API_KEY)


def prepare_staging(client: Client, index_name: str, searchable: list[str], filterable: list[str], sortable: list[str]):
    staging = index_name + "_staging"
    try:
        client.delete_index(staging)
        time.sleep(0.5)
    except Exception:
        pass
    client.create_index(staging, {"primaryKey": "id"})
    time.sleep(1.5)
    idx = client.index(staging)
    idx.update_settings({
        "searchableAttributes": searchable,
        "filterableAttributes": filterable,
        "sortableAttributes": sortable,
    })
    time.sleep(1)
    try:
        client.index(index_name).get_stats()
    except Exception:
        log(f"Indexul live {index_name} nu există. Îl creez...")
        client.create_index(index_name, {"primaryKey": "id"})
        time.sleep(1)
        client.index(index_name).update_settings({
            "searchableAttributes": searchable,
            "filterableAttributes": filterable,
            "sortableAttributes": sortable,
        })
        time.sleep(1)


def index_docs(client: Client, index_name: str, docs: list[dict], dry_run: bool):
    if not docs:
        log(f"  {index_name}: 0 documente — nimic de indexat")
        return 0
    prepare_staging(client, index_name,
                    _searchable_for(index_name),
                    _filterable_for(index_name),
                    _sortable_for(index_name))
    staging = index_name + "_staging"
    target = client.index(staging)
    t0 = time.time()
    for i in range(0, len(docs), BATCH_SIZE):
        batch = docs[i:i + BATCH_SIZE]
        if not dry_run:
            task = target.add_documents(batch)
            try:
                client.wait_for_task(task.task_uid, timeout_in_ms=30000)
            except Exception:
                time.sleep(2)
    log(f"  {index_name}: {len(docs)} docs indexate în {time.time() - t0:.1f}s")
    return len(docs)


def _searchable_for(index_name: str) -> list[str]:
    if index_name == INDEX_ABSORBTIE:
        return ["program"]
    return ["titlu", "beneficiar", "judet", "regiune", "stadiu", "program"]


def _filterable_for(index_name: str) -> list[str]:
    if index_name == INDEX_ABSORBTIE:
        return ["perioada"]
    return ["program", "judet", "stadiu"]


def _sortable_for(index_name: str) -> list[str]:
    if index_name == INDEX_ABSORBTIE:
        return []
    return ["valoare_totala", "valoare_eligibila", "plati"]


def swap_indexes(client: Client, index_name: str):
    staging = index_name + "_staging"
    client.swap_indexes([{"indexes": [index_name, staging]}])
    time.sleep(1)
    try:
        client.delete_index(staging)
    except Exception:
        pass
    log(f"Swap atomic complet: {staging} → {index_name}")


# ── Main ───────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--max", type=int, default=None, help="Max programe proiecte + resurse absorbție de procesat")
    parser.add_argument("--force", action="store_true", help="Ignoră state-ul (re-procesează tot)")
    args = parser.parse_args()

    state = load_state()

    # ── 1. Proiecte contractate ─────────────────
    log("Descopăr pachetul 'Proiecte contractate'...")
    pkg = find_projects_package()
    if pkg:
        resources = latest_resource_per_program(pkg)
        log(f"  {len(resources)} programe cu resurse recente: {', '.join(sorted(resources.keys()))}")

        # State hash: {program: report_date} — skip dacă neschimbat (fără --force)
        cur_hash = json.dumps({p: resources[p].get("_report_date", "") for p in resources}, sort_keys=True)
        if not args.force and state.get("mfe_projects_hash") == cur_hash:
            log("  State neschimbat (aceleași date de raportare) — sar peste proiecte.")
        else:
            all_docs = []
            for prog in sorted(resources.keys()):
                res = resources[prog]
                if args.max and len(all_docs) > 0 and sum(1 for _ in [prog]) >= args.max:
                    pass
                log(f"  Descarc {prog} ({res.get('name', '')[:50]})...")
                try:
                    resp = requests.get(res["url"], headers=HEADERS, timeout=120)
                    if resp.status_code != 200 or len(resp.content) < 1000:
                        log(f"    HTTP {resp.status_code} / gol — sar peste")
                        continue
                except Exception as e:
                    log(f"    Eroare: {e} — sar peste")
                    continue
                docs = parse_projects(resp.content, prog, res.get("_report_date", ""))
                log(f"    {len(docs)} proiecte")
                all_docs.extend(docs)
                time.sleep(0.6)
                if args.max and len(all_docs) >= args.max * 100:
                    break

            log(f"Total proiecte: {len(all_docs)}")
            if all_docs:
                client = get_ms_client()
                index_docs(client, INDEX_NAME, all_docs, args.dry_run)
                if not args.dry_run:
                    swap_indexes(client, INDEX_NAME)
                    state["mfe_projects_hash"] = cur_hash
                    state["mfe_projects_count"] = len(all_docs)
    else:
        log("  Pachet 'Proiecte contractate' negăsit!")

    # ── 2. Stadiul absorbției ───────────────────
    log("Descopăr pachetele de absorbție...")
    absorb_pkgs = find_absorb_packages()
    if absorb_pkgs:
        all_absorb = []
        for pkg in absorb_pkgs:
            resources = sorted(pkg.get("resources", []), key=lambda r: r.get("created", "") or "", reverse=True)
            if not resources:
                continue
            res = resources[0]  # cea mai recentă
            perioada = re.search(r"(\d{1,2}\s+\w+\s+\d{4})", res.get("name", ""))
            perioada_s = perioada.group(1) if perioada else res.get("name", "")[:20]
            log(f"  Descarc absorbție: {res.get('name', '')[:55]}...")
            try:
                resp = requests.get(res["url"], headers=HEADERS, timeout=120)
                if resp.status_code != 200 or len(resp.content) < 1000:
                    log(f"    HTTP {resp.status_code} / gol — sar peste")
                    continue
            except Exception as e:
                log(f"    Eroare: {e} — sar peste")
                continue
            docs = parse_absorb(resp.content, perioada_s)
            log(f"    {len(docs)} programe absorbție")
            all_absorb.extend(docs)
            time.sleep(0.6)

        log(f"Total absorbție: {len(all_absorb)}")
        if all_absorb:
            client = get_ms_client()
            index_docs(client, INDEX_ABSORBTIE, all_absorb, args.dry_run)
            if not args.dry_run:
                swap_indexes(client, INDEX_ABSORBTIE)
                state["mfe_absorb_hash"] = datetime.now().strftime("%Y%m%d")
                state["mfe_absorb_count"] = len(all_absorb)
    else:
        log("  Pachete de absorbție negăsite!")

    if not args.dry_run:
        save_state(state)
        log("State salvat")
    else:
        log("Dry-run: nu am indexat nimic")


if __name__ == "__main__":
    main()
