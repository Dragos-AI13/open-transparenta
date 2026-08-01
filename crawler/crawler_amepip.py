#!/usr/bin/env python3
"""
Crawler AMEPIP — întreprinderi publice (OUG 109/2011).

Descarcă datele AMEPIP (Agenția pentru Monitorizarea și Evaluarea
Performanțelor Întreprinderilor Publice) de pe data.gov.ro și le indexează
în Meilisearch (index "intreprinderi_publice", upsert pe CUI).

Sursa principală: data_2023.csv — curat, 1.259 firme, ani 2019-2023,
17 indicatori financiari specifici (ROE, ROA, EBITDA, lichiditate, marje).

Sursa opțională: datecompanii_ind-finnefin.xlsx (ian 2026) — aduce anul 2024,
format pivot cu coduri interne AMEPIP; se încearcă maparea, altfel se sare.

Usage:
    python crawler_amepip.py                   # Indexează tot (CSV 2023)
    python crawler_amepip.py --include-2024    # Încearcă și XLSX 2024
    python crawler_amepip.py --max=10          # Doar primele 10 firme
    python crawler_amepip.py --dry-run         # Doar descarcă, nu indexează
    python crawler_amepip.py --force           # Forțează re-descărcare
"""

import argparse
import csv
import hashlib
import json
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path

import requests
from dotenv import load_dotenv
from meilisearch import Client

# ── Config ─────────────────────────────────────

load_dotenv(Path(__file__).resolve().parent.parent / "frontend" / ".env")

MEILISEARCH_HOST = os.getenv("MEILISEARCH_HOST", "http://localhost:7700")
MEILISEARCH_API_KEY = os.getenv("MEILISEARCH_API_KEY", "")
INDEX_NAME = "intreprinderi_publice"
BATCH_SIZE = 100
STATE_FILE = Path(__file__).parent / ".crawler_state.json"
CACHE_CSV = Path(__file__).parent / ".crawler_amepip.csv"

CKAN_BASE = "https://data.gov.ro"

# Indicatorii din CSV care ne interesează (cheie canonicală)
CSV_INDICATORI = [
    "ROE",
    "ROA",
    "Rata de crestere a profitului net",
    "Rata de crestere a cifrei de afaceri",
    "Marja de profit net",
    "Marja de profit din exploatare",
    "Viteza de rotatie a stocurilor",
    "Viteza de rotatie a creantelor",
    "Viteza de rotatie a activelor",
    "Datorie vs EBITDA",
    "Datorii totale",
    "EBITDA",
    "Levierul",
    "Lichiditate imediată",
    "Rata lichiditatii Curente",
    "Cota de piata",
]


def log(msg: str):
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] {msg}", flush=True)


# ── CKAN helpers ───────────────────────────────

def find_amepip_csv_resource() -> dict | None:
    """Găsește resursa CSV cu indicatori financiari AMEPIP (2023)."""
    url = f"{CKAN_BASE}/api/3/action/package_search"
    params = {"q": "AMEPIP", "sort": "metadata_modified desc", "rows": 10}
    resp = requests.get(url, params=params, timeout=30)
    resp.raise_for_status()
    for pkg in resp.json().get("result", {}).get("results", []):
        for res in pkg.get("resources", []):
            name = (res.get("name") or "").upper()
            if "INDICATORI FINANCIARI" in name and res.get("format") == "CSV":
                res["_dataset_title"] = pkg.get("title", "")
                return res
    return None


def find_amepip_xlsx_resource() -> dict | None:
    """Găsește resursa XLSX ian 2026 (indicatori + guvernanță, an 2024)."""
    url = f"{CKAN_BASE}/api/3/action/package_search"
    params = {"q": "AMEPIP", "sort": "metadata_modified desc", "rows": 10}
    resp = requests.get(url, params=params, timeout=30)
    resp.raise_for_status()
    for pkg in resp.json().get("result", {}).get("results", []):
        for res in pkg.get("resources", []):
            name = (res.get("name") or "").upper()
            if "DATE FINANCIARE SI NON-FINANCIARE" in name and res.get("format") == "XLSX":
                res["_dataset_title"] = pkg.get("title", "")
                return res
    return None


def get_download_url(resource_id: str) -> str:
    url = f"{CKAN_BASE}/api/3/action/resource_show?id={resource_id}"
    resp = requests.get(url, timeout=30)
    resp.raise_for_status()
    return resp.json()["result"]["url"]


# ── State ──────────────────────────────────────

def load_state() -> dict:
    if STATE_FILE.exists():
        try:
            return json.loads(STATE_FILE.read_text())
        except Exception:
            pass
    return {"amepip_last_hash": None, "amepip_last_count": 0}


def save_state(state: dict):
    STATE_FILE.write_text(json.dumps(state, indent=2, default=str))


# ── Parsing ────────────────────────────────────

def parse_val(raw: str) -> float | None:
    """Parsează valori de tip '10,74%', '30556079', '-1,83%' etc."""
    if raw is None:
        return None
    s = str(raw).strip().replace("%", "").replace(" ", "").strip()
    if not s or s in ("-", "--", "n/a", "N/A"):
        return None
    s = s.replace(",", ".").replace("−", "-")
    try:
        return float(s)
    except ValueError:
        return None


def normalize_cui(raw: str) -> str:
    """54760 → RO54760. Doar cifre după RO (fără zero-pad)."""
    digits = re.sub(r"\D", "", str(raw or ""))
    if not digits:
        return ""
    return f"RO{digits}"


def split_caen(raw: str) -> tuple[str, str]:
    """'3600-Captarea, tratarea si distributia apei' → ('3600', 'Captarea...')."""
    if not raw:
        return ("", "")
    s = str(raw).strip()
    m = re.match(r"^(\d+)\s*[-–]\s*(.*)$", s)
    if m:
        return (m.group(1), m.group(2).strip())
    m2 = re.match(r"^(\d+)\s*(.*)$", s)
    if m2:
        return (m2.group(1), m2.group(2).strip())
    return ("", s)


# ── XLSX 2024 (opțional) ───────────────────────

def try_parse_xlsx_2024(xlsx_path: Path) -> dict[str, dict]:
    """
    Încearcă să extragă datele din XLSX (an 2024).
    Format pivot cu coduri interne — dacă nu găsim mapare CUI, returnăm {}.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        log("openpyxl nu e instalat — sar peste XLSX 2024 (pip install openpyxl)")
        return {}

    log("Parsez XLSX 2024 (pivot)...")
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as e:
        log(f"XLSX 2024: nu pot deschide fișierul ({e}) — sar peste")
        return {}

    # Căutăm un sheet cu coloane numite (name/cui) — încercăm toate sheet-urile
    for ws in wb.worksheets:
        # header row
        headers: list[str] = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(c).strip() if c is not None else "" for c in row]
            break
        low = [h.lower() for h in headers]
        if "name" not in low or "an" not in low:
            continue
        name_idx = low.index("name")
        an_idx = low.index("an")
        cui_idx = low.index("cui") if "cui" in low else None

        out: dict[str, dict] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) <= max(name_idx, an_idx):
                continue
            name = str(row[name_idx] or "").strip()
            an = row[an_idx]
            cui = normalize_cui(row[cui_idx]) if cui_idx is not None and row[cui_idx] else ""
            if not name or an is None:
                continue
            # Mapăm coloanele rămase ca indicatori (nume coloană → valoare)
            indicators = {}
            for i, h in enumerate(headers):
                if i in (name_idx, an_idx, cui_idx) or i >= len(row):
                    continue
                v = parse_val(row[i])
                if v is not None:
                    indicators[h] = v
            firm = out.setdefault(cui or name, {"denumire": name, "ani_indicatori": {}})
            firm["denumire"] = name
            firm["ani_indicatori"][int(an)] = indicators
        if out:
            log(f"XLSX 2024: am extras {len(out)} firme din sheet '{ws.title}'")
            return out

    log("XLSX 2024: nu am găsit structură utilizabilă — sar peste (folosesc doar CSV 2023)")
    return {}


# ── Meilisearch ────────────────────────────────

STAGING_INDEX = INDEX_NAME + "_staging"


def get_ms_client() -> Client:
    return Client(url=MEILISEARCH_HOST, api_key=MEILISEARCH_API_KEY)


def prepare_staging_index(client: Client):
    try:
        client.delete_index(STAGING_INDEX)
        time.sleep(0.5)
    except Exception:
        pass

    client.create_index(STAGING_INDEX, {"primaryKey": "cui"})
    time.sleep(1.5)

    idx = client.index(STAGING_INDEX)
    idx.update_settings({
        "searchableAttributes": ["denumire", "cui", "caen_denumire"],
        "filterableAttributes": ["caen", "ani"],
        "sortableAttributes": ["denumire"],
    })
    time.sleep(1)

    try:
        client.index(INDEX_NAME).get_stats()
    except Exception:
        log("Indexul live nu există. Îl creez...")
        client.create_index(INDEX_NAME, {"primaryKey": "cui"})
        time.sleep(1)
        client.index(INDEX_NAME).update_settings({
            "searchableAttributes": ["denumire", "cui", "caen_denumire"],
            "filterableAttributes": ["caen", "ani"],
            "sortableAttributes": ["denumire"],
        })
        time.sleep(1)


def swap_indexes(client: Client):
    client.swap_indexes([{"indexes": [INDEX_NAME, STAGING_INDEX]}])
    time.sleep(1)
    try:
        client.delete_index(STAGING_INDEX)
    except Exception:
        pass
    log("Swap atomic complet: staging → live")


# ── Main ───────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--max", type=int, default=None)
    parser.add_argument("--include-2024", action="store_true")
    parser.add_argument("--force", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    state = load_state()

    # 1. Sursa principală: CSV
    log("Caut resursa CSV AMEPIP pe data.gov.ro...")
    csv_res = find_amepip_csv_resource()
    if not csv_res:
        log("Eroare: nu am găsit resursa CSV AMEPIP")
        sys.exit(1)
    log(f"CSV: {csv_res.get('name')} ({csv_res.get('size', '?')} B) — {csv_res.get('_dataset_title')}")

    csv_id = csv_res.get("id", "")
    csv_hash = hashlib.sha256(f"amepip_csv_{csv_id}".encode()).hexdigest()[:16]

    if state.get("amepip_last_hash") == csv_hash and not args.force and CACHE_CSV.exists():
        log("CSV-ul e deja la zi (hash identic). Folosește --force pentru re-descărcare.")
    else:
        url = get_download_url(csv_id)
        log(f"Descarc {url}")
        try:
            with requests.get(url, stream=True, timeout=60) as resp:
                resp.raise_for_status()
                with open(CACHE_CSV, "wb") as f:
                    for chunk in resp.iter_content(chunk_size=1024 * 256):
                        f.write(chunk)
            log("Descărcare completă")
        except Exception as e:
            log(f"Eroare la descărcare: {e}")
            if not CACHE_CSV.exists():
                sys.exit(1)

    # 2. (Opțional) XLSX 2024
    xlsx_2024: dict[str, dict] = {}
    if args.include_2024:
        xlsx_res = find_amepip_xlsx_resource()
        if xlsx_res:
            log(f"XLSX: {xlsx_res.get('name')}")
            xlsx_path = Path(__file__).parent / ".crawler_amepip_2024.xlsx"
            if not xlsx_path.exists() or args.force:
                try:
                    url = get_download_url(xlsx_res.get("id", ""))
                    with requests.get(url, stream=True, timeout=120) as resp:
                        resp.raise_for_status()
                        with open(xlsx_path, "wb") as f:
                            for chunk in resp.iter_content(chunk_size=1024 * 256):
                                f.write(chunk)
                    log(f"XLSX descărcat ({xlsx_path.stat().st_size / 1024 / 1024:.1f} MB)")
                except Exception as e:
                    log(f"XLSX: eroare la descărcare ({e}) — sar peste")
            if xlsx_path.exists():
                xlsx_2024 = try_parse_xlsx_2024(xlsx_path)
        else:
            log("XLSX 2024: resursa nu a fost găsită — sar peste")

    # 3. Agregăm CSV-ul pe firme
    log("Agregare CSV pe firme (CUI → indicatori pe ani)...")
    firms: dict[str, dict] = {}
    with open(CACHE_CSV, "r", encoding="utf-8-sig", errors="replace") as f:
        reader = csv.DictReader(f)
        for i, row in enumerate(reader):
            if args.max and len(firms) >= args.max:
                break

            cui = normalize_cui(row.get("cui"))
            name = (row.get("name") or "").strip()
            if not cui or not name:
                continue

            an_raw = row.get("an") or ""
            try:
                an = int(float(an_raw))
            except ValueError:
                continue

            firm = firms.setdefault(cui, {
                "cui": cui,
                "denumire": name.upper(),
                "numar_registru_comert": (row.get("registration_number") or "").strip(),
                "ticker_symbol": (row.get("ticker_symbol") or "").strip(),
                "caen": "",
                "caen_denumire": "",
                "ani": [],
                "indicatori": {},
            })

            caen_code, caen_desc = split_caen(row.get("CAEN") or "")
            if caen_code and not firm["caen"]:
                firm["caen"] = caen_code
                firm["caen_denumire"] = caen_desc

            if an not in firm["ani"]:
                firm["ani"].append(an)

            ind_an = firm["indicatori"].setdefault(an, {})
            for key in CSV_INDICATORI:
                v = parse_val(row.get(key))
                if v is not None:
                    ind_an[key] = v

    # 4. Îmbinăm XLSX 2024 (dacă avem)
    if xlsx_2024:
        merged_2024 = 0
        for cui_x, data_x in xlsx_2024.items():
            for cui_f, firm in firms.items():
                if cui_f == cui_x or (not cui_x.startswith("RO") and cui_f.endswith(cui_x)):
                    for an, inds in data_x.get("ani_indicatori", {}).items():
                        if an not in firm["ani"]:
                            firm["ani"].append(an)
                        firm["indicatori"].setdefault(an, {}).update(inds)
                    merged_2024 += 1
                    break
        log(f"XLSX 2024: îmbinate date pentru {merged_2024} firme")

    for firm in firms.values():
        firm["ani"].sort()

    log(f"Total firme unice: {len(firms)}")
    if not firms:
        log("Nicio firmă procesată — ies.")
        sys.exit(1)

    # 5. Indexăm (staging + swap atomic)
    client = get_ms_client()
    prepare_staging_index(client)
    target = client.index(STAGING_INDEX)

    docs = list(firms.values())
    t0 = time.time()
    for i in range(0, len(docs), BATCH_SIZE):
        batch = docs[i:i + BATCH_SIZE]
        if not args.dry_run:
            target.add_documents(batch)
        done = min(i + BATCH_SIZE, len(docs))
        if done % 300 == 0 or done == len(docs):
            log(f"  {done}/{len(docs)} indexate")
    elapsed = time.time() - t0
    log(f"Gata: {len(docs)} firme indexate în {elapsed:.0f}s")

    if not args.dry_run and docs:
        swap_indexes(client)
        state["amepip_last_hash"] = csv_hash
        state["amepip_last_count"] = len(docs)
        save_state(state)
        log("State salvat")
    elif args.dry_run:
        log("Dry-run: nu am indexat nimic")


if __name__ == "__main__":
    main()
