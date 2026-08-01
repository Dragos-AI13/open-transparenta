#!/usr/bin/env python3
"""
Crawler ME — Rezultate Bacalaureat (rate de promovare pe școli/județe).

Descarcă de pe data.gov.ro pachetul „Rezultate Bacalaureat {sesiune}"
(org: ministerul-educatiei) + rețeaua școlară pentru join pe SIIIR,
agreghează pe școală și indexează în Meilisearch.

Structură bacalaureat (verificat 2026-08-01 pe sesiunea 2-2025, 7MB):
    r1: header, 52 coloane — Cod unic candidat, Sex, Specializare, Profil,
        Fileira, Forma de învățământ, Mediu candidat, Unitate (SIIIR),
        Unitate (SIRUES), Clasa, Promoție, NOTE_RECUN_A..D, STATUS_A..ED,
        NOTA_EA..ED, CONTESTATIE_*, PUNCTAJ DIGITALE, STATUS, Medie
    r2+: candidați (30.280) — ANONIMIZAȚI (cod unic, fără nume)
    STATUS: „Promovat" | „Nepromovat" | „Absent" | „Eliminat"
    Medie: col52 (numeric, 1-10)

Agregare: pe școală (SIIIR) → {judet (din rețea), denumire școală,
candidati, promovati, rata_promovare}. Join-ul se face LOCAL (parsam și
rețeaua școlară) — codurile SIIIR nu se pot pagina complet prin API
Meilisearch (cap 1000).

Usage:
    python crawler_educatie_bac.py              # Indexează tot
    python crawler_educatie_bac.py --dry-run    # Doar descarcă+parsează
    python crawler_educatie_bac.py --force      # Ignoră state-ul
"""

import argparse
import json
import os
import re
import sys
import time
from datetime import datetime
from io import BytesIO
from pathlib import Path

import openpyxl
import requests
from dotenv import load_dotenv
from meilisearch import Client

# ── Config ─────────────────────────────────────

load_dotenv(Path(__file__).resolve().parent.parent / "frontend" / ".env")

MEILISEARCH_HOST = os.getenv("MEILISEARCH_HOST", "http://localhost:7700")
MEILISEARCH_API_KEY = os.getenv("MEILISEARCH_API_KEY", "")
INDEX_NAME = "bacalaureat"
BATCH_SIZE = 500
STATE_FILE = Path(__file__).parent / ".crawler_state.json"

CKAN_API = "https://data.gov.ro/api/3/action/package_search"
HEADERS = {
    "User-Agent": "OpenTransparentaBot/1.0 (+https://github.com/Dragos-AI13/open-transparenta; date publice ME)",
    "Accept": "application/json",
}

# Reutilizăm parserul rețelei pentru join
from crawler_educatie_retea import parse_retea, find_retea_package, latest_xlsx as latest_xlsx_retea

LUNI_RO = {
    "ianuarie": 1, "februarie": 2, "martie": 3, "aprilie": 4, "mai": 5,
    "iunie": 6, "iulie": 7, "august": 8, "septembrie": 9, "octombrie": 10,
    "noiembrie": 11, "decembrie": 12,
}


def log(msg: str):
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] {msg}", flush=True)


# ── State ──────────────────────────────────────

def load_state() -> dict:
    if STATE_FILE.exists():
        try:
            return json.loads(STATE_FILE.read_text())
        except Exception:
            pass
    return {"bac_hash": None, "bac_count": 0}


def save_state(state: dict):
    STATE_FILE.write_text(json.dumps(state, indent=2, default=str))


# ── CKAN discovery ─────────────────────────────

def parse_report_date(name: str) -> tuple:
    """(an, lună, zi) din numele pachetului — comparație numerică.
    „sesiunea 2- 2025" nu are lună → fallback pe an (2025, 0, 0)."""
    m = re.search(
        r"(\d{1,2})\s+(ianuarie|februarie|martie|aprilie|mai|iunie|iulie|august|septembrie|octombrie|noiembrie|decembrie)\s+(\d{4})",
        name, re.I,
    )
    if m:
        return (int(m.group(3)), LUNI_RO[m.group(2).lower()], int(m.group(1)))
    # Fallback: anul de la sfârșit („sesiunea 2- 2025", „2024 sesiunea 1")
    m2 = re.search(r"(\d{4})", name)
    if m2:
        return (int(m2.group(1)), 0, 0)
    return (0, 0, 0)


def find_bac_package() -> dict | None:
    """Cel mai recent pachet „Rezultate Bacalaureat" (org ministerul-educatiei)."""
    best, best_date = None, (0, 0, 0)
    try:
        resp = requests.get(
            CKAN_API,
            params={"q": "Rezultate Bacalaureat", "rows": 15, "sort": "metadata_modified desc"},
            headers=HEADERS,
            timeout=30,
        )
        resp.raise_for_status()
        for pkg in resp.json()["result"]["results"]:
            org = (pkg.get("organization") or {}).get("name", "")
            title = pkg.get("title", "")
            if org == "ministerul-educatiei" and "bacalaureat" in title.lower():
                d = parse_report_date(title)
                if d > best_date:
                    best, best_date = pkg, d
    except Exception as e:
        log(f"  CKAN error: {e}")
    return best


def latest_xlsx(pkg: dict) -> dict | None:
    resources = [
        r for r in pkg.get("resources", [])
        if "xlsx" in (r.get("format") or "").lower()
        or (r.get("name") or "").lower().endswith(".xlsx")
    ]
    if not resources:
        return None
    return sorted(resources, key=lambda r: r.get("created", "") or "", reverse=True)[0]


# ── Parsing bacalaureat ────────────────────────

def parse_bac(content: bytes) -> list[dict]:
    """
    Candidați: {cod, sex, specializare, profil, fileira, forma, mediu,
    siiir, promovat: bool}. PERFORMANȚĂ: iter_rows, nu ws.cell.
    """
    wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    hdr = [str(v or "").strip() for v in rows[0]]
    col = {}
    for i, h in enumerate(hdr):
        if h == "Cod unic candidat":
            col["cod"] = i
        elif h == "Sex":
            col["sex"] = i
        elif h == "Specializare":
            col["specializare"] = i
        elif h == "Profil":
            col["profil"] = i
        elif h == "Mediu candidat":
            col["mediu"] = i
        elif h == "Unitate (SIIIR)":
            col["siiir"] = i
        elif h == "STATUS":
            col["status"] = i

    if "status" not in col:
        return []

    cand = []
    for row in rows[1:]:
        cod = row[col["cod"]] if "cod" in col else None
        siiir = str(row[col["siiir"]]).strip() if row[col["siiir"]] is not None else ""
        status = str(row[col["status"]] or "").strip()
        cand.append({
            "cod": str(cod) if cod is not None else "",
            "sex": str(row[col["sex"]] or "").strip() if "sex" in col else "",
            "specializare": str(row[col["specializare"]] or "").strip() if "specializare" in col else "",
            "profil": str(row[col["profil"]] or "").strip() if "profil" in col else "",
            "mediu": str(row[col["mediu"]] or "").strip().upper() if "mediu" in col else "",
            "siiir": siiir,
            "promovat": status == "Promovat",
            "prezent": status in ("Promovat", "Nepromovat"),
        })
    return cand


# ── Agregare ───────────────────────────────────

def aggregate_by_school(cand: list[dict], siiir_map: dict[str, dict]) -> list[dict]:
    """Agreghează pe școală (SIIIR): candidați, promovați, rată + județ din rețea."""
    schools: dict[str, dict] = {}
    for c in cand:
        if not c["siiir"]:
            continue
        s = schools.setdefault(c["siiir"], {
            "siiir": c["siiir"],
            "candidati": 0,
            "prezenti": 0,
            "promovati": 0,
        })
        s["candidati"] += 1
        if c["prezent"]:
            s["prezenti"] += 1
        if c["promovat"]:
            s["promovati"] += 1

    docs = []
    for siiir, s in schools.items():
        meta = siiir_map.get(siiir, {})
        docs.append({
            "id": f"bac_{siiir}",
            "siiir": siiir,
            "denumire": meta.get("denumire", ""),
            "judet": meta.get("judet", ""),
            "judet_nume": meta.get("judet_nume", ""),
            "localitate": meta.get("localitate", ""),
            "candidati": s["candidati"],
            "prezenti": s["prezenti"],
            "promovati": s["promovati"],
            "rata_promovare": round((s["promovati"] / s["prezenti"]) * 100, 2) if s["prezenti"] else 0.0,
            "sesiune": "sesiunea 2-2025",
        })
    return docs


# ── Meilisearch ────────────────────────────────

STAGING_INDEX = INDEX_NAME + "_staging"


def get_ms_client() -> Client:
    return Client(url=MEILISEARCH_HOST, api_key=MEILISEARCH_API_KEY)


def prepare_staging_index(client: Client):
    try:
        client.delete_index(STAGING_INDEX)
        time.sleep(0.5)
    except Exception:
        pass
    client.create_index(STAGING_INDEX, {"primaryKey": "id"})
    time.sleep(1.5)
    idx = client.index(STAGING_INDEX)
    idx.update_settings({
        "searchableAttributes": ["denumire", "judet_nume", "localitate"],
        "filterableAttributes": ["judet", "judet_nume", "sesiune"],
        "sortableAttributes": ["rata_promovare", "candidati"],
    })
    time.sleep(1)
    try:
        client.index(INDEX_NAME).get_stats()
    except Exception:
        log(f"Indexul live {INDEX_NAME} nu există. Îl creez...")
        client.create_index(INDEX_NAME, {"primaryKey": "id"})
        time.sleep(1)
        client.index(INDEX_NAME).update_settings({
            "searchableAttributes": ["denumire", "judet_nume", "localitate"],
            "filterableAttributes": ["judet", "judet_nume", "sesiune"],
            "sortableAttributes": ["rata_promovare", "candidati"],
        })
        time.sleep(1)


def swap_indexes(client: Client):
    client.swap_indexes([{"indexes": [INDEX_NAME, STAGING_INDEX]}])
    time.sleep(1)
    try:
        client.delete_index(STAGING_INDEX)
    except Exception:
        pass
    log("Swap atomic complet: staging → live")


# ── Main ───────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--force", action="store_true")
    args = parser.parse_args()

    state = load_state()

    log("Descopăr pachetul 'Rezultate Bacalaureat'...")
    pkg = find_bac_package()
    if not pkg:
        log("Pachet negăsit!")
        sys.exit(1)
    res = latest_xlsx(pkg)
    if not res:
        log("Nicio resursă XLSX!")
        sys.exit(1)
    log(f"  Pachet: {pkg.get('title', '')[:60]}")
    log(f"  Resursă: {res.get('name', '')[:60]}")

    cur_hash = pkg.get("title", "") + res.get("created", "")
    if not args.force and state.get("bac_hash") == cur_hash:
        log("State neschimbat — sar peste.")
        return

    # 1. Descarcă bacalaureatul
    log("Descarc bacalaureat XLSX...")
    try:
        resp = requests.get(res["url"], headers=HEADERS, timeout=(10, 180))
        if resp.status_code != 200 or len(resp.content) < 1000:
            log(f"  HTTP {resp.status_code} / gol — ies.")
            sys.exit(1)
    except Exception as e:
        log(f"  Eroare descărcare: {e}")
        sys.exit(1)

    cand = parse_bac(resp.content)
    log(f"  Candidați: {len(cand)}")

    # 2. Descarcă rețeaua școlară pentru join SIIIR → județ/denumire
    log("Descarc rețeaua școlară (join SIIIR)...")
    siiir_map: dict[str, dict] = {}
    try:
        retea_pkg = find_retea_package()
        retea_res = latest_xlsx_retea(retea_pkg) if retea_pkg else None
        if retea_res:
            r2 = requests.get(retea_res["url"], headers=HEADERS, timeout=(10, 180))
            if r2.status_code == 200 and len(r2.content) > 1000:
                for d in parse_retea(r2.content):
                    if d.get("cod_siiir"):
                        siiir_map[d["cod_siiir"]] = d
                log(f"  Rețea: {len(siiir_map)} coduri SIIIR mapate")
    except Exception as e:
        log(f"  Eroare rețea: {e} (continuăm fără județ)")

    # 3. Agregare pe școală
    docs = aggregate_by_school(cand, siiir_map)
    log(f"  Școli agregate: {len(docs)}")
    if docs:
        cu_judet = sum(1 for d in docs if d["judet_nume"])
        log(f"  Cu județ din rețea: {cu_judet} / {len(docs)}")

    if not docs:
        log("Nicio dată agregată. Ies.")
        sys.exit(1)

    if args.dry_run:
        log("Dry-run: nu am indexat.")
        for d in sorted(docs, key=lambda x: -x["rata_promovare"])[:5]:
            log(f"  {d['denumire'][:40]} | {d['judet_nume']} | cand {d['candidati']} | rata {d['rata_promovare']}%")
        return

    client = get_ms_client()
    prepare_staging_index(client)
    target = client.index(STAGING_INDEX)
    t0 = time.time()
    for i in range(0, len(docs), BATCH_SIZE):
        batch = docs[i:i + BATCH_SIZE]
        task = target.add_documents(batch)
        try:
            client.wait_for_task(task.task_uid, timeout_in_ms=60000)
        except Exception:
            time.sleep(2)
    log(f"Indexat în {time.time() - t0:.1f}s")

    swap_indexes(client)
    state["bac_hash"] = cur_hash
    state["bac_count"] = len(docs)
    save_state(state)
    log("State salvat")


if __name__ == "__main__":
    main()
