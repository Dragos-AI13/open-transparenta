#!/usr/bin/env python3
"""
Crawler ME — Rețea școlară (toate unitățile de învățământ din România).

Descarcă de pe data.gov.ro pachetul „Rețea scolară {an școlar}" (org:
ministerul-educatiei) și indexează unitățile în Meilisearch.

Structură (verificat 2026-08-01 pe fișier real, 3MB, an școlar 2025-2026):
    r1: „Generat la data: ..."
    r2-r3: gol
    r4: header (An, Judet PJ, Localitate PJ, Cod SIRUTA PJ, Mediu loc. PJ,
        Cod SIIIR PJ, Denumire PJ, Localitate unitate, Mediu loc. unitate,
        Cod SIRUES, Cod SIIIR unitate, Denumire scurta unitate,
        Denumire lunga unitate, Tip unitate, Statut unitate, Cod fiscal,
        Mod functionare, Forma finantare, Forma proprietate, Strada, Numar,
        Cod postal, Telefon, Fax, Email)
    r5+: date (18.026 rânduri)
    Aceeași unitate apare de mai multe ori (r5 = r6 identice) → id unic
    pe Cod SIIIR unitate + hash(denumire|localitate|strada).

Usage:
    python crawler_educatie_retea.py              # Indexează tot
    python crawler_educatie_retea.py --dry-run    # Doar descarcă+parsează
    python crawler_educatie_retea.py --max=500    # Doar primele 500
    python crawler_educatie_retea.py --force      # Ignoră state-ul
"""

import argparse
import hashlib
import json
import os
import re
import sys
import time
from datetime import datetime
from io import BytesIO
from pathlib import Path

import openpyxl
import requests
from dotenv import load_dotenv
from meilisearch import Client

# ── Config ─────────────────────────────────────

load_dotenv(Path(__file__).resolve().parent.parent / "frontend" / ".env")

MEILISEARCH_HOST = os.getenv("MEILISEARCH_HOST", "http://localhost:7700")
MEILISEARCH_API_KEY = os.getenv("MEILISEARCH_API_KEY", "")
INDEX_NAME = "retea_scolara"
BATCH_SIZE = 500
STATE_FILE = Path(__file__).parent / ".crawler_state.json"

CKAN_API = "https://data.gov.ro/api/3/action/package_search"
HEADERS = {
    "User-Agent": "OpenTransparentaBot/1.0 (+https://github.com/Dragos-AI13/open-transparenta; date publice ME)",
    "Accept": "application/json",
}

# Coduri județ → nume (pentru afișare și filtrare)
JUDETE_RO = {
    "AB": "Alba", "AG": "Argeș", "AR": "Arad", "B": "București",
    "BC": "Bacău", "BH": "Bihor", "BN": "Bistrița-Năsăud", "BR": "Brăila",
    "BT": "Botoșani", "BV": "Brașov", "BZ": "Buzău", "CJ": "Cluj",
    "CL": "Călărași", "CS": "Caraș-Severin", "CT": "Constanța",
    "CV": "Covasna", "DB": "Dâmbovița", "DJ": "Dolj", "GJ": "Gorj",
    "GL": "Galați", "GR": "Giurgiu", "HD": "Hunedoara", "HR": "Harghita",
    "IF": "Ilfov", "IL": "Ialomița", "IS": "Iași", "MH": "Mehedinți",
    "MM": "Maramureș", "MS": "Mureș", "NT": "Neamț", "OT": "Olt",
    "PH": "Prahova", "SB": "Sibiu", "SJ": "Sălaj", "SM": "Satu Mare",
    "SV": "Suceava", "TL": "Tulcea", "TM": "Timiș", "TR": "Teleorman",
    "VL": "Vâlcea", "VN": "Vrancea", "VS": "Vaslui",
}


def log(msg: str):
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] {msg}", flush=True)


# ── Utilitare ───────────────────────────────────

def normalize_ro(s: str) -> str:
    return (
        s.replace("\u015f", "\u0219")
        .replace("\u0163", "\u021b")
        .replace("\u015e", "\u0218")
        .replace("\u0162", "\u021a")
    )


def doc_hash(*parts: str) -> str:
    return hashlib.sha1("|".join(parts).encode()).hexdigest()[:16]


# ── State ──────────────────────────────────────

def load_state() -> dict:
    if STATE_FILE.exists():
        try:
            return json.loads(STATE_FILE.read_text())
        except Exception:
            pass
    return {"retea_hash": None, "retea_count": 0}


def save_state(state: dict):
    STATE_FILE.write_text(json.dumps(state, indent=2, default=str))


# ── CKAN discovery ─────────────────────────────

def find_retea_package() -> dict | None:
    """Cel mai recent pachet „Rețea scolară" (org ministerul-educatiei)."""
    try:
        resp = requests.get(
            CKAN_API,
            params={
                "q": "Rețea scolară",
                "rows": 10,
                "sort": "metadata_modified desc",
            },
            headers=HEADERS,
            timeout=30,
        )
        resp.raise_for_status()
        for pkg in resp.json()["result"]["results"]:
            org = (pkg.get("organization") or {}).get("name", "")
            title = pkg.get("title", "")
            if org == "ministerul-educatiei" and "rețea" in title.lower():
                return pkg
    except Exception as e:
        log(f"  CKAN error: {e}")
    return None


def latest_xlsx(pkg: dict) -> dict | None:
    """Cea mai recentă resursă XLSX din pachet (după created)."""
    resources = [
        r for r in pkg.get("resources", [])
        if "xlsx" in (r.get("format") or "").lower()
        or (r.get("name") or "").lower().endswith(".xlsx")
    ]
    if not resources:
        return None
    return sorted(resources, key=lambda r: r.get("created", "") or "", reverse=True)[0]


# ── Parsing ────────────────────────────────────

def find_header_row(ws) -> int:
    """Header-ul: primul rând care conține „Judet PJ" și „Denumire"."""
    for r in range(1, 10):
        vals = [str(ws.cell(row=r, column=c).value or "") for c in range(1, 27)]
        joined = " ".join(vals)
        if "judet" in joined.lower() and ("denumire" in joined.lower()):
            return r
    return 4  # fallback verificat


def map_columns(ws, hdr: int) -> dict[str, int]:
    """Mapare coloane după nume — normalizare ASCII pentru matching."""
    col: dict[str, int] = {}
    for c in range(1, 27):
        v = ws.cell(row=hdr, column=c).value
        if not v:
            continue
        key = str(v).strip().lower()
        key = (
            key.replace("ă", "a").replace("â", "a").replace("î", "i")
            .replace("ș", "s").replace("ş", "s").replace("ț", "t").replace("ţ", "t")
        )
        if "an" == key or key.startswith("an "):
            col["an"] = c
        elif "judet" in key:
            col["judet"] = c
        elif "localitate" in key and "unitate" not in key:
            col["localitate"] = c
        elif "mediu loc" in key and "unitate" not in key:
            col["mediu"] = c
        elif "cod siiir" in key and "unitate" in key:
            col["cod_siiir"] = c
        elif "denumire scurta unitate" in key:
            col["denumire_scurta"] = c
        elif "denumire lunga unitate" in key:
            col["denumire_lunga"] = c
        elif "tip unitate" in key:
            col["tip"] = c
        elif "statut unitate" in key:
            col["statut"] = c
        elif "mod functionare" in key:
            col["mod_functionare"] = c
        elif "forma finantare" in key:
            col["finantare"] = c
        elif "strada" in key:
            col["strada"] = c
        elif "numar" == key:
            col["numar"] = c
        elif "cod postal" in key:
            col["cod_postal"] = c
        elif "telefon" in key:
            col["telefon"] = c
        elif "email" in key:
            col["email"] = c
    return col


def parse_retea(content: bytes) -> list[dict]:
    """
    Parsează rețeaua școlară → documente Meilisearch.

    PERFORMANȚĂ (pitfall): `ws.cell(row, col)` pe un workbook read_only e
    FOARTE lent (~15s/500 rânduri). Folosim `iter_rows(values_only=True)`
    care întoarce toate celulele unui rând dintr-o dată (~0.1s/1000).
    """
    wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Citim TOATE rândurile o singură dată (listă de tupluri)
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Header: primul rând care conține „Judet PJ" + „Denumire"
    hdr_idx = None
    for i, row in enumerate(rows[:12]):
        joined = " ".join(str(v or "") for v in row[:27]).lower()
        if "judet" in joined and "denumire" in joined:
            hdr_idx = i
            break
    if hdr_idx is None:
        hdr_idx = 3  # fallback verificat (r4 → index 3)

    # Mapare coloane după nume (normalizare ASCII)
    col: dict[str, int] = {}
    for c, v in enumerate(rows[hdr_idx][:27]):
        if not v:
            continue
        key = str(v).strip().lower()
        key = (
            key.replace("ă", "a").replace("â", "a").replace("î", "i")
            .replace("ș", "s").replace("ş", "s").replace("ț", "t").replace("ţ", "t")
        )
        if key == "an" or key.startswith("an "):
            col["an"] = c
        elif "judet" in key:
            col["judet"] = c
        elif "localitate" in key and "unitate" not in key:
            col["localitate"] = c
        elif "mediu loc" in key and "unitate" not in key:
            col["mediu"] = c
        elif "cod siiir" in key and "unitate" in key:
            col["cod_siiir"] = c
        elif "denumire scurta unitate" in key:
            col["denumire_scurta"] = c
        elif "denumire lunga unitate" in key:
            col["denumire_lunga"] = c
        elif "tip unitate" in key:
            col["tip"] = c
        elif "statut unitate" in key:
            col["statut"] = c
        elif "mod functionare" in key:
            col["mod_functionare"] = c
        elif "forma finantare" in key:
            col["finantare"] = c
        elif "strada" in key:
            col["strada"] = c
        elif key == "numar":
            col["numar"] = c
        elif "cod postal" in key:
            col["cod_postal"] = c
        elif "telefon" in key:
            col["telefon"] = c
        elif "email" in key:
            col["email"] = c

    if "denumire_scurta" not in col and "denumire_lunga" not in col:
        return []

    def g(row, name):
        c = col.get(name)
        return row[c] if c is not None and c < len(row) else None

    docs = []
    for row in rows[hdr_idx + 1:]:
        den_scurta = g(row, "denumire_scurta")
        den_lunga = g(row, "denumire_lunga")
        if not den_scurta and not den_lunga:
            continue
        denumire = normalize_ro(str(den_scurta or den_lunga).strip())
        if not denumire:
            continue

        judet_cod = str(g(row, "judet") or "").strip().upper()
        localitate = normalize_ro(str(g(row, "localitate") or "").strip())
        mediu = str(g(row, "mediu") or "").strip().upper()
        cod_siiir = str(g(row, "cod_siiir") or "").strip()
        tip = normalize_ro(str(g(row, "tip") or "").strip())
        statut = normalize_ro(str(g(row, "statut") or "").strip())
        mod_func = normalize_ro(str(g(row, "mod_functionare") or "").strip())
        finantare = normalize_ro(str(g(row, "finantare") or "").strip())
        strada = normalize_ro(str(g(row, "strada") or "").strip())
        numar = str(g(row, "numar") or "").strip()
        cod_postal = str(g(row, "cod_postal") or "").strip()
        telefon = str(g(row, "telefon") or "").strip()
        email = str(g(row, "email") or "").strip()

        adresa = ", ".join(x for x in [strada, numar] if x)
        judet_nume = JUDETE_RO.get(judet_cod, judet_cod)

        # id: {cod_siiir}_{hash} — aceeași unitate apare de mai multe ori
        pid = f"{cod_siiir}_{doc_hash(denumire, localitate, strada)}" if cod_siiir else f"n{doc_hash(denumire, localitate, strada)}"

        docs.append({
            "id": pid,
            "denumire": denumire,
            "judet": judet_cod,
            "judet_nume": judet_nume,
            "localitate": localitate or None,
            "mediu": mediu or None,
            "cod_siiir": cod_siiir or None,
            "tip": tip or None,
            "statut": statut or None,
            "mod_functionare": mod_func or None,
            "finantare": finantare or None,
            "adresa": adresa or None,
            "cod_postal": cod_postal or None,
            "telefon": telefon or None,
            "email": email or None,
            "an": "2025-2026",
        })
    return docs


# ── Meilisearch ────────────────────────────────

STAGING_INDEX = INDEX_NAME + "_staging"


def get_ms_client() -> Client:
    return Client(url=MEILISEARCH_HOST, api_key=MEILISEARCH_API_KEY)


def prepare_staging_index(client: Client):
    try:
        client.delete_index(STAGING_INDEX)
        time.sleep(0.5)
    except Exception:
        pass
    client.create_index(STAGING_INDEX, {"primaryKey": "id"})
    time.sleep(1.5)
    idx = client.index(STAGING_INDEX)
    idx.update_settings({
        "searchableAttributes": ["denumire", "localitate", "judet_nume", "email"],
        "filterableAttributes": ["judet", "judet_nume", "mediu", "tip"],
        "sortableAttributes": ["denumire"],
    })
    time.sleep(1)
    try:
        client.index(INDEX_NAME).get_stats()
    except Exception:
        log(f"Indexul live {INDEX_NAME} nu există. Îl creez...")
        client.create_index(INDEX_NAME, {"primaryKey": "id"})
        time.sleep(1)
        client.index(INDEX_NAME).update_settings({
            "searchableAttributes": ["denumire", "localitate", "judet_nume", "email"],
            "filterableAttributes": ["judet", "judet_nume", "mediu", "tip"],
            "sortableAttributes": ["denumire"],
        })
        time.sleep(1)


def swap_indexes(client: Client):
    client.swap_indexes([{"indexes": [INDEX_NAME, STAGING_INDEX]}])
    time.sleep(1)
    try:
        client.delete_index(STAGING_INDEX)
    except Exception:
        pass
    log("Swap atomic complet: staging → live")


# ── Main ───────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--max", type=int, default=None)
    parser.add_argument("--force", action="store_true")
    args = parser.parse_args()

    state = load_state()

    log("Descopăr pachetul 'Rețea scolară'...")
    pkg = find_retea_package()
    if not pkg:
        log("Pachet negăsit!")
        sys.exit(1)

    res = latest_xlsx(pkg)
    if not res:
        log("Nicio resursă XLSX în pachet!")
        sys.exit(1)
    log(f"  Resursă: {res.get('name', '')[:70]}")

    # State hash pe numele resursei — skip dacă neschimbat
    cur_hash = res.get("name", "") + res.get("created", "")
    if not args.force and state.get("retea_hash") == cur_hash:
        log("State neschimbat (aceeași resursă) — sar peste.")
        return

    log("Descarc XLSX...")
    try:
        resp = requests.get(res["url"], headers=HEADERS, timeout=120)
        if resp.status_code != 200 or len(resp.content) < 1000:
            log(f"  HTTP {resp.status_code} / gol — ies.")
            sys.exit(1)
    except Exception as e:
        log(f"  Eroare: {e}")
        sys.exit(1)

    docs = parse_retea(resp.content)
    log(f"Parse: {len(docs)} unități")
    if args.max:
        docs = docs[: args.max]
        log(f"  (limitat la {args.max})")

    if not docs:
        log("Nicio dată extrasă. Ies.")
        sys.exit(1)

    if args.dry_run:
        log("Dry-run: nu am indexat nimic.")
        return

    client = get_ms_client()
    prepare_staging_index(client)
    target = client.index(STAGING_INDEX)

    t0 = time.time()
    for i in range(0, len(docs), BATCH_SIZE):
        batch = docs[i:i + BATCH_SIZE]
        task = target.add_documents(batch)
        try:
            client.wait_for_task(task.task_uid, timeout_in_ms=60000)
        except Exception:
            time.sleep(2)
    log(f"Indexat în {time.time() - t0:.1f}s")

    swap_indexes(client)
    state["retea_hash"] = cur_hash
    state["retea_count"] = len(docs)
    save_state(state)
    log("State salvat")


if __name__ == "__main__":
    main()
